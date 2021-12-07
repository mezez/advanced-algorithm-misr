from types import new_class
from openpyxl import *
import pandas as pd
from scipy.spatial import distance_matrix


def read_tsp_file(file=None):
    if type(file) is list:
        final_data = [file, len(file[0])]
    else:
        if file is None:
            file = "./tsp_database/bays297by7.xlsx"
        wb = load_workbook(file)
        sheet = wb.worksheets[0]
        row_count = sheet.max_row
        #if uploaded file is a list of coordinates, then convert them to cost matrices
        #otherwise assume uploaded files are already matrices

        if sheet.max_column == 2:
            #file contains cordinates, generate matrix from coordinates
            coordinates = [[0 for x in range(sheet.max_column)]
                           for y in range(row_count)]

            for i in range(row_count):
                for j in range(sheet.max_column):
                    coordinates[i][j] = sheet.cell(i + 1, j + 1).value
            ctys = [0 for x in range(row_count)]
            for i in range(row_count):
                ctys[i] = i
            df = pd.DataFrame(coordinates, columns=['xcord', 'ycord'], index=ctys)
            cost_matrix = distance_matrix(df.values, df.values)
            print(cost_matrix)

        else:
            column_count = sheet.max_column
            cost_matrix = [[0 for x in range(column_count)]
                           for y in range(row_count)]
            for i in range(column_count):
                for j in range(row_count):
                    cost_matrix[i][j] = sheet.cell(i + 1, j + 1).value
        final_data = [cost_matrix, row_count]
    return final_data


def convertPath(path):
    newPath = ""
    for i in range(len(path)):
        if i == len(path) - 1:
            newPath += str(int(path[i]))
        else:
            newPath += str(int(path[i])) + " -> "
    return newPath
